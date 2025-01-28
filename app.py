
from flask import Flask, request, jsonify,  render_template, request, redirect, url_for, session
from flask_cors import CORS
import os
import pyodbc
import bcrypt
import pulp
import math
import threading
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from flask import send_file


static_folder = os.path.join(os.getcwd(), 'static')
if not os.path.exists(static_folder):
    os.makedirs(static_folder)


app = Flask(__name__)
CORS(app)  # Enable Cross-Origin Resource Sharing (CORS)

# Define your connection string
conn_str = (
    r'DRIVER={ODBC Driver 17 for SQL Server};'
    r'SERVER=YOUR-COMPUTER\SQLEXPRESS;'  # Your server instance
    r'DATABASE=CANN;'  # Your database name
    r'Trusted_Connection=yes;'  # Use Windows Authentication
)
conn = pyodbc.connect(conn_str)
app.secret_key = "1"


@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data['username']
    password = data['password']
    phone = data['phone']

    try:
        # Connect to the database
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # SQL to insert user data
        query = "INSERT INTO Users (username, password, phone) VALUES (?, ?, ?)"
        cursor.execute(query, (username, password, phone))
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({'message': 'Registration successful!'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Global variable to store the logged-in user's ID
logged_in_user_id = None

@app.route('/login', methods=['POST'])
def login():
    global logged_in_user_id  # Access the global variable

    data = request.get_json()
    username = data['username']
    password = data['password']

    try:
        # Connect to the database
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Query to find the user by username
        query = "SELECT userID, username, password FROM Users WHERE username = ?"
        cursor.execute(query, (username,))
        user = cursor.fetchone()

        # Check if the user exists and if the password matches
        if user and user.password == password:
            # Store the userID in the global variable
            logged_in_user_id = user.userID

            # Optionally store in session as well
            session['userID'] = logged_in_user_id

            return jsonify({'message': 'Login successful!', 'userID': logged_in_user_id}), 200
        else:
            return jsonify({'error': 'Invalid username or password'}), 401

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()



@app.route('/get_guests', methods=['GET'])
def get_guests():
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()


        user_id = logged_in_user_id

        query = "SELECT * FROM Guest WHERE userID = ?"
        cursor.execute(query, (user_id,))
        guests = cursor.fetchall()

        guest_list = []
        for guest in guests:
            guest_list.append({
                'guestID': guest.guestID,
                'guest_name': guest.guest_name,
                'guest_category': guest.guest_category,
                'phone': guest.phone
            })

        conn.close()
        return jsonify(guest_list)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route to add a guest
@app.route('/add_guest', methods=['POST'])
def add_guest():
    try:
        data = request.get_json()
        guest_name = data['guest_name']
        guest_category = data['guest_category']
        phone = data['phone']


        user_id = logged_in_user_id

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Insert the guest into the database
        query = "INSERT INTO Guest (guest_name, guest_category, phone, userID) VALUES (?, ?, ?, ?)"
        cursor.execute(query, (guest_name, guest_category, phone, user_id))

        # Commit the transaction
        conn.commit()
        conn.close()

        return jsonify({'message': 'Guest added successfully!'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route to remove a guest
@app.route('/remove_guest', methods=['DELETE'])
def remove_guest():
    try:
        data = request.get_json()
        guest_id = data['guestID']

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Delete the guest from the database
        query = "DELETE FROM Guest WHERE guestID = ?"
        cursor.execute(query, (guest_id,))

        # Commit the transaction
        conn.commit()
        conn.close()

        return jsonify({'message': 'Guest removed successfully!'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500




# Function to create a thin border style
def create_thin_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Function to create a thick border style
def create_thick_border():
    return Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

def set_column_widths(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

class SolverThread(threading.Thread):
    def __init__(self, prob, timeout):
        threading.Thread.__init__(self)
        self.prob = prob
        self.timeout = timeout
        self.status = None

    def run(self):
        self.prob.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=self.timeout))
        self.status = pulp.LpStatus[self.prob.status]


#create the seating plan
@app.route('/run_seating', methods=['POST'])
def run_seating():
    try:
        # Debugging: Print the request data received from the frontend
        data = request.get_json()
        print("Received data from frontend:", data)  # Log the received data

        max_guests_per_table = data.get('max_guests_per_table')

        if not max_guests_per_table or max_guests_per_table <= 0:
            return jsonify({'error': 'Please provide a valid number of guests per table.'}), 400

        # Establish the database connection
        conn = pyodbc.connect(conn_str)

        # Use pd.read_sql to fetch guest data
        query = "SELECT guest_name, guest_category FROM Guest WHERE userID = ?"
        guest_data = pd.read_sql(query, conn, params=(logged_in_user_id,))

        total_guests = len(guest_data)
        categories = guest_data['guest_category'].unique()
        tables = math.ceil(total_guests / max_guests_per_table)

        prob = pulp.LpProblem("Wedding_Seating", pulp.LpMinimize)

        # Define binary variables for guest assignments and category assignments
        x = pulp.LpVariable.dicts("guest", (range(total_guests), range(tables)), cat='Binary')
        y = pulp.LpVariable.dicts("category", (categories, range(tables)), cat='Binary')

        prob += pulp.lpSum([y[c][j] for c in categories for j in range(tables)])

        # Ensure each guest is assigned to exactly one table
        for i in range(total_guests):
            prob += pulp.lpSum([x[i][j] for j in range(tables)]) == 1

        # Ensure no table exceeds the guest limit
        for j in range(tables):
            prob += pulp.lpSum([x[i][j] for i in range(total_guests)]) <= max_guests_per_table

        # Link category assignment to guests
        for c in categories:
            for j in range(tables):
                for i in range(total_guests):
                    if guest_data.iloc[i]['guest_category'] == c:
                        prob += y[c][j] >= x[i][j]

        # Set a timeout for solving the problem
        time_limit = 120
        solver_thread = SolverThread(prob, time_limit)
        solver_thread.start()
        solver_thread.join(time_limit)

        if solver_thread.is_alive():
            print("Solver time limit exceeded. Stopping the solver.")
            solver_thread.join()

        print(f"Status: {solver_thread.status}")

        # Prepare the table assignments
        table_data = {f"Table {j + 1}": [] for j in range(tables)}
        guest_names = guest_data['guest_name'].tolist()

        for j in range(tables):
            for i in range(total_guests):
                if pulp.value(x[i][j]) == 1:
                    table_data[f"Table {j + 1}"].append(guest_names[i])

        # שמירת סידור ההושבה למסד הנתונים
        try:
            save_seating_plan_internal(table_data)
        except Exception as e:
            print(f"Error saving seating plan to database: {str(e)}")

        # Create the seating plan in an Excel file
        output_file = os.path.join('static', 'seating_plan.xlsx')
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            table_assignments = pd.DataFrame.from_dict(table_data, orient='index').transpose()
            table_assignments.to_excel(writer, sheet_name="Seating Plan", index=False)

        # Send the seating data to the frontend
        return jsonify({
            "message": "Seating plan generated successfully!",
            "download_link": f"/static/seating_plan.xlsx",
            "table_data": table_data  # Add the seating data here
        })

    except Exception as e:
        print(f"Error: {str(e)}")  # Log any error to the server log
        return jsonify({'error': str(e)}), 500



#save seating plan
def save_seating_plan_internal(table_data):
    user_id = logged_in_user_id
    if not user_id:
        raise Exception('User not logged in')

    # התחברות למסד הנתונים
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # מחיקת סידור קודם של המשתמש
    delete_query = "DELETE FROM seating_plan WHERE user_id = ?"
    cursor.execute(delete_query, (user_id,))

    # שמירת הסידור החדש
    for table_number, guests in table_data.items():
        insert_query = "INSERT INTO seating_plan (user_id, table_number, guests) VALUES (?, ?, ?)"
        cursor.execute(insert_query, (user_id, table_number, ', '.join(guests)))

    conn.commit()
    conn.close()



#שליפת סידור ההושבה מהמסד
@app.route('/get_seating_plan', methods=['GET'])
def get_seating_plan():
    try:
        user_id = logged_in_user_id
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401

        # Connect to the database
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Retrieve seating plan for the user
        query = "SELECT table_number, guests FROM seating_plan WHERE user_id = ?"
        cursor.execute(query, (user_id,))
        rows = cursor.fetchall()

        table_data = {row.table_number: row.guests.split(', ') for row in rows}

        conn.close()
        return jsonify({'table_data': table_data}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@app.route('/download_seating_plan', methods=['GET'])
def download_seating_plan():
    try:
        user_id = logged_in_user_id
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401

        # Connect to the database
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Retrieve seating plan for the user
        query = "SELECT table_number, guests FROM seating_plan WHERE user_id = ?"
        cursor.execute(query, (user_id,))
        rows = cursor.fetchall()
        table_data = {row.table_number: row.guests.split(', ') for row in rows}

        # Retrieve guest categories
        category_query = "SELECT guest_name, guest_category FROM guest"
        cursor.execute(category_query)
        category_rows = cursor.fetchall()
        guest_categories = {row.guest_name: row.guest_category for row in category_rows}

        # Calculate categories count and the objective value
        categories_count = []
        for table, guests in table_data.items():
            unique_categories = set(guest_categories.get(guest, "") for guest in guests)
            categories_count.append(len(unique_categories))

        # Calculate the objective value as the sum of all unique categories counts
        objective_value = sum(categories_count)

        # Create the Excel file
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Seating Plan"

        # Helper functions for styling
        def create_thin_border():
            return Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        def create_thick_border():
            return Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )

        def set_column_widths(worksheet):
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Define styles
        thick_border = create_thick_border()
        thin_border = create_thin_border()

        # Main header
        worksheet['A1'] = 'סידור הושבה'
        worksheet['A1'].font = Font(size=14, bold=True, color='000000')
        worksheet['A1'].fill = PatternFill(start_color='D8BBAB', end_color='D8BBAB', fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A1'].border = thin_border
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(table_data))

        # Table headers (table numbers as columns)
        column = 1
        for table in table_data.keys():
            cell = worksheet.cell(row=2, column=column, value=f"שולחן {table}")
            cell.font = Font(bold=True, color='000000')
            cell.fill = PatternFill(start_color='ECE1DA', end_color='ECE1DA', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            column += 1

        # Write guest data below each table number
        max_rows = 0  # Track the maximum number of guests for any table
        for col_index, (table, guests) in enumerate(table_data.items(), start=1):
            for row_index, guest in enumerate(guests, start=3):
                worksheet.cell(row=row_index, column=col_index, value=guest).border = thin_border
            max_rows = max(max_rows, len(guests))

        # Define the fill color for light gray (D9D9D9)
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Add categories count row
        categories_row = max_rows + 4
        worksheet[f'A{categories_row}'] = 'סכום קטגוריות לשולחן'
        worksheet[f'A{categories_row}'].font = Font(bold=True)
        worksheet[f'A{categories_row}'].alignment = Alignment(horizontal='center')
        worksheet[f'A{categories_row}'].border = thin_border
        worksheet[f'A{categories_row}'].fill = light_gray_fill  # Apply the fill color
        worksheet.merge_cells(start_row=categories_row, start_column=1, end_row=categories_row,
                              end_column=len(table_data))

        # Write category counts for each table in individual cells under each table header
        for col_index, value in enumerate(categories_count, start=1):
            cell = worksheet.cell(row=categories_row + 1, column=col_index, value=value)
            cell.border = thin_border
            cell.fill = light_gray_fill  # Apply the fill color

        # Add objective value row
        objective_row = categories_row + 3
        worksheet[f'A{objective_row}'] = 'ערך פונקציית מטרה'
        worksheet[f'A{objective_row}'].font = Font(bold=True)
        worksheet[f'A{objective_row}'].alignment = Alignment(horizontal='center')
        worksheet[f'A{objective_row}'].border = thin_border
        worksheet[f'A{objective_row}'].fill = light_gray_fill  # Apply the fill color
        worksheet.merge_cells(start_row=objective_row, start_column=1, end_row=objective_row, end_column=2)

        # Assign the objective value to the third column, which is outside the merged range
        worksheet[f'C{objective_row}'] = objective_value
        worksheet[f'C{objective_row}'].border = thin_border
        worksheet[f'C{objective_row}'].fill = light_gray_fill  # Apply the fill color

        # Adjust column widths
        set_column_widths(worksheet)

        # Save the workbook
        workbook.save(output)
        output.seek(0)

        conn.close()

        return send_file(output, download_name="seating_plan.xlsx", as_attachment=True)

    except Exception as e:
        return jsonify({'error': str(e)}), 500



if __name__ == '__main__':
    app.run(debug=True)
